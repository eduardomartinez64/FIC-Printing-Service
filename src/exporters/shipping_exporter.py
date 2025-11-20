"""
Excel Exporter for Shopify Shipping Data.
Exports shipping profiles with zones, countries, and rates in a hierarchical,
easy-to-read format for logistics operations review.
"""

import logging
from datetime import datetime
from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ShippingExporter:
    """Exports Shopify shipping data to Excel with one sheet per profile."""

    def __init__(self):
        """Initialize the exporter."""
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet

        # Define styles
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Zone header style (for zone names)
        self.zone_font = Font(bold=True, size=12, color="FFFFFF")
        self.zone_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Subheader style (for "Countries Served", "Shipping Rates" sections)
        self.subheader_font = Font(bold=True, size=11, color="FFFFFF")
        self.subheader_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

        # Rate header style
        self.rate_header_font = Font(bold=True, size=10)
        self.rate_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Border style
        self.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

    def _sanitize_sheet_name(self, name: str, max_length: int = 31) -> str:
        """
        Sanitize sheet name to be valid for Excel.
        Excel sheet names must be <= 31 chars and cannot contain: / \\ ? * [ ]
        """
        # Replace invalid characters
        invalid_chars = ['/', '\\', '?', '*', '[', ']', ':']
        for char in invalid_chars:
            name = name.replace(char, '-')

        # Truncate if too long
        if len(name) > max_length:
            name = name[:max_length]

        return name

    def _auto_adjust_columns(self, sheet, min_width: int = 12, max_width: int = 60):
        """Auto-adjust column widths based on content."""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        # Handle wrapped text - split by newlines
                        lines = str(cell.value).split('\n')
                        max_line_length = max(len(line) for line in lines)
                        max_length = max(max_length, max_line_length)
                except:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _format_rate_description(self, rate: Dict, rate_type: str) -> str:
        """
        Create a human-readable description of a shipping rate rule.

        Args:
            rate: Rate dictionary
            rate_type: Type of rate (weight, price, or carrier)

        Returns:
            Formatted description string
        """
        if rate_type == "weight":
            weight_low = rate.get('weight_low', 0)
            weight_high = rate.get('weight_high', 'unlimited')

            # Format weight values
            if weight_high == 'unlimited' or weight_high is None or weight_high == '':
                return f"Weight: {weight_low}+ lbs"
            else:
                return f"Weight: {weight_low} - {weight_high} lbs"

        elif rate_type == "price":
            min_price = rate.get('min_order_subtotal')
            max_price = rate.get('max_order_subtotal')

            # Format price values
            if min_price is None and max_price is None:
                return "All order values"
            elif max_price is None or max_price == '':
                return f"Order value: ${min_price}+"
            elif min_price is None or min_price == '':
                return f"Order value: up to ${max_price}"
            else:
                return f"Order value: ${min_price} - ${max_price}"

        elif rate_type == "carrier":
            service_filter = rate.get('service_filter', {})
            carrier_id = rate.get('carrier_service_id', 'Unknown')

            if service_filter.get('*'):
                return f"Carrier: All services (ID: {carrier_id})"
            else:
                services = ', '.join(service_filter.values()) if service_filter else 'All'
                return f"Carrier: {services} (ID: {carrier_id})"

        return "Standard rate"

    def export_to_excel(self, shipping_data: Dict[str, Any], output_file: str) -> str:
        """
        Export shipping data to Excel file with one sheet per profile.

        Args:
            shipping_data: Dictionary containing profiles and related data
            output_file: Output file path

        Returns:
            Path to created Excel file
        """
        logger.info(f"Starting export to: {output_file}")

        try:
            # Create overview sheet
            self._create_overview_sheet(shipping_data)

            # Create one sheet per profile
            profiles = shipping_data.get('profiles', [])

            if not profiles:
                # Fallback: If no profiles, create from zones
                logger.warning("No profiles found, creating single sheet from zones")
                zones = shipping_data.get('shipping_zones', [])
                self._create_profile_sheet({
                    'name': 'All Shipping Zones',
                    'id': 'default',
                    'zones': zones
                }, 1)
            else:
                for idx, profile in enumerate(profiles, 1):
                    self._create_profile_sheet(profile, idx)

            # Save workbook
            self.workbook.save(output_file)
            logger.info(f"Successfully exported data to: {output_file}")
            return output_file

        except Exception as e:
            logger.error(f"Failed to export data: {e}")
            raise

    def _create_overview_sheet(self, shipping_data: Dict[str, Any]):
        """Create overview/summary sheet."""
        sheet = self.workbook.create_sheet("Overview", 0)

        # Add title
        sheet['A1'] = "Shopify Shipping Configuration Export"
        sheet['A1'].font = Font(size=16, bold=True, color="366092")
        sheet.merge_cells('A1:D1')

        # Add metadata
        row = 3
        metadata = [
            ("Export Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("", ""),
            ("Configuration Summary:", ""),
        ]

        for label, value in metadata:
            sheet[f'A{row}'] = label
            sheet[f'B{row}'] = value
            if label:
                sheet[f'A{row}'].font = Font(bold=True)
            row += 1

        # Count statistics
        profiles = shipping_data.get('profiles', [])
        total_zones = sum(len(p.get('zones', [])) for p in profiles)
        total_rates = 0

        for profile in profiles:
            for zone in profile.get('zones', []):
                total_rates += len(zone.get('weight_based_shipping_rates', []))
                total_rates += len(zone.get('price_based_shipping_rates', []))
                total_rates += len(zone.get('carrier_shipping_rate_providers', []))

        stats = [
            ("Shipping Profiles:", len(profiles)),
            ("Total Shipping Zones:", total_zones),
            ("Total Shipping Rates:", total_rates),
            ("Carrier Services:", len(shipping_data.get('carrier_services', []))),
        ]

        for label, value in stats:
            sheet[f'A{row}'] = label
            sheet[f'B{row}'] = value
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1

        # Add profile breakdown
        row += 2
        sheet[f'A{row}'] = "Shipping Profiles:"
        sheet[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        # Header for profile list
        sheet[f'A{row}'] = "Profile Name"
        sheet[f'B{row}'] = "Zones"
        sheet[f'C{row}'] = "Rates"
        sheet[f'D{row}'] = "Sheet Name"

        for col in ['A', 'B', 'C', 'D']:
            sheet[f'{col}{row}'].font = self.rate_header_font
            sheet[f'{col}{row}'].fill = self.rate_header_fill
        row += 1

        # List each profile
        for profile in profiles:
            profile_name = profile.get('name', f"Profile {profile.get('id', 'Unknown')}")
            zones = profile.get('zones', [])

            # Count rates in this profile
            profile_rates = 0
            for zone in zones:
                profile_rates += len(zone.get('weight_based_shipping_rates', []))
                profile_rates += len(zone.get('price_based_shipping_rates', []))
                profile_rates += len(zone.get('carrier_shipping_rate_providers', []))

            sheet_name = self._sanitize_sheet_name(profile_name)

            sheet[f'A{row}'] = profile_name
            sheet[f'B{row}'] = len(zones)
            sheet[f'C{row}'] = profile_rates
            sheet[f'D{row}'] = sheet_name
            row += 1

        # Add navigation instructions
        row += 2
        sheet[f'A{row}'] = "📋 How to Use This Export:"
        sheet[f'A{row}'].font = Font(bold=True, size=11)
        row += 1

        instructions = [
            "• Each shipping profile has its own tab/sheet",
            "• Within each profile, zones are listed with their countries and rates",
            "• Rates include: Rate Name, Price, and Description/Rule",
            "• Use this data to identify consolidation opportunities",
            "• Look for duplicate rate structures across zones",
        ]

        for instruction in instructions:
            sheet[f'A{row}'] = instruction
            row += 1

        self._auto_adjust_columns(sheet)

    def _create_profile_sheet(self, profile: Dict, index: int):
        """
        Create a sheet for a single shipping profile with hierarchical layout.

        Args:
            profile: Profile dictionary containing zones
            index: Profile index for unique sheet naming
        """
        profile_name = profile.get('name', f"Profile {profile.get('id', index)}")
        sheet_name = self._sanitize_sheet_name(profile_name)

        # Ensure unique sheet names
        if sheet_name in self.workbook.sheetnames:
            sheet_name = self._sanitize_sheet_name(f"{profile_name}_{index}")

        sheet = self.workbook.create_sheet(sheet_name)
        logger.info(f"Creating sheet: {sheet_name}")

        row = 1

        # Profile title
        sheet[f'A{row}'] = profile_name
        sheet[f'A{row}'].font = Font(size=14, bold=True, color="366092")
        sheet.merge_cells(f'A{row}:D{row}')
        row += 2

        # Get zones for this profile
        zones = profile.get('zones', [])

        if not zones:
            sheet[f'A{row}'] = "No shipping zones configured for this profile"
            sheet[f'A{row}'].font = Font(italic=True, color="999999")
            self._auto_adjust_columns(sheet)
            return

        # Process each zone
        for zone in zones:
            zone_name = zone.get('name', 'Unnamed Zone')

            # Zone header (merged across columns)
            sheet[f'A{row}'] = f"📍 {zone_name}"
            sheet.merge_cells(f'A{row}:D{row}')
            sheet[f'A{row}'].font = self.zone_font
            sheet[f'A{row}'].fill = self.zone_fill
            sheet[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
            row += 1

            # Countries served section
            countries = zone.get('countries', [])
            if countries:
                sheet[f'A{row}'] = "Countries Served:"
                sheet[f'A{row}'].font = self.subheader_font
                sheet[f'A{row}'].fill = self.subheader_fill
                sheet.merge_cells(f'A{row}:D{row}')
                row += 1

                # List countries
                country_list = []
                for country in countries:
                    country_name = country.get('name', country.get('code', 'Unknown'))
                    provinces = country.get('provinces', [])

                    if provinces:
                        # List specific provinces
                        province_names = [p.get('name', p.get('code', '')) for p in provinces]
                        country_list.append(f"{country_name} ({', '.join(province_names[:5])}{'...' if len(province_names) > 5 else ''})")
                    else:
                        # Entire country
                        country_list.append(f"{country_name} (All regions)")

                # Write countries (chunk into readable groups)
                for i in range(0, len(country_list), 3):
                    chunk = country_list[i:i+3]
                    sheet[f'A{row}'] = "  • " + ", ".join(chunk)
                    sheet.merge_cells(f'A{row}:D{row}')
                    row += 1

                row += 1  # Spacing

            # Shipping rates section
            weight_rates = zone.get('weight_based_shipping_rates', [])
            price_rates = zone.get('price_based_shipping_rates', [])
            carrier_rates = zone.get('carrier_shipping_rate_providers', [])

            total_rates = len(weight_rates) + len(price_rates) + len(carrier_rates)

            if total_rates > 0:
                sheet[f'A{row}'] = f"Shipping Rates ({total_rates} total):"
                sheet[f'A{row}'].font = self.subheader_font
                sheet[f'A{row}'].fill = self.subheader_fill
                sheet.merge_cells(f'A{row}:D{row}')
                row += 1

                # Rate table headers
                sheet[f'A{row}'] = "Rate Name"
                sheet[f'B{row}'] = "Price"
                sheet[f'C{row}'] = "Type"
                sheet[f'D{row}'] = "Rule / Description"

                for col in ['A', 'B', 'C', 'D']:
                    sheet[f'{col}{row}'].font = self.rate_header_font
                    sheet[f'{col}{row}'].fill = self.rate_header_fill
                    sheet[f'{col}{row}'].border = self.border
                    sheet[f'{col}{row}'].alignment = Alignment(horizontal="center")
                row += 1

                # Weight-based rates
                for rate in weight_rates:
                    sheet[f'A{row}'] = rate.get('name', 'Unnamed Rate')
                    sheet[f'B{row}'] = f"${rate.get('price', '0.00')}"
                    sheet[f'C{row}'] = "Weight-Based"
                    sheet[f'D{row}'] = self._format_rate_description(rate, 'weight')

                    for col in ['A', 'B', 'C', 'D']:
                        sheet[f'{col}{row}'].border = self.border

                    sheet[f'B{row}'].alignment = Alignment(horizontal="right")
                    row += 1

                # Price-based rates
                for rate in price_rates:
                    sheet[f'A{row}'] = rate.get('name', 'Unnamed Rate')
                    sheet[f'B{row}'] = f"${rate.get('price', '0.00')}"
                    sheet[f'C{row}'] = "Price-Based"
                    sheet[f'D{row}'] = self._format_rate_description(rate, 'price')

                    for col in ['A', 'B', 'C', 'D']:
                        sheet[f'{col}{row}'].border = self.border

                    sheet[f'B{row}'].alignment = Alignment(horizontal="right")
                    row += 1

                # Carrier rates
                for rate in carrier_rates:
                    flat_rate = rate.get('flat_modifier_type_id')
                    carrier_service = rate.get('carrier_service_id', 'Unknown')

                    rate_name = f"Carrier Service {carrier_service}"
                    if flat_rate:
                        rate_name += " (Flat Rate)"

                    sheet[f'A{row}'] = rate_name
                    sheet[f'B{row}'] = "Calculated"
                    sheet[f'C{row}'] = "Carrier Service"
                    sheet[f'D{row}'] = self._format_rate_description(rate, 'carrier')

                    for col in ['A', 'B', 'C', 'D']:
                        sheet[f'{col}{row}'].border = self.border

                    sheet[f'B{row}'].alignment = Alignment(horizontal="right")
                    row += 1

            else:
                sheet[f'A{row}'] = "No shipping rates configured"
                sheet[f'A{row}'].font = Font(italic=True, color="999999")
                row += 1

            # Add spacing between zones
            row += 2

        # Auto-adjust column widths
        self._auto_adjust_columns(sheet, min_width=15)

        # Set specific column widths for better readability
        sheet.column_dimensions['A'].width = 35  # Rate Name
        sheet.column_dimensions['B'].width = 12  # Price
        sheet.column_dimensions['C'].width = 18  # Type
        sheet.column_dimensions['D'].width = 50  # Description

        logger.info(f"Completed sheet '{sheet_name}' with {len(zones)} zones")
